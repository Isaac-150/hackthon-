from flask import Flask, render_template, request, jsonify, send_file
import openpyxl
from openpyxl import Workbook
import os
import datetime
import re
import json
from uuid import uuid4

app = Flask(__name__)

# Initialize data files
def init_data_files():
    if not os.path.exists('data'):
        os.makedirs('data')
    
    # Product data
    product_filepath = 'data/artisandata.xlsx'
    if not os.path.exists(product_filepath):
        wb = Workbook()
        ws = wb.active
        ws.append(['ID', 'Product Name', 'Category', 'Description_EN', 'Description_TA', 'Quantity', 'Price', 'Timestamp'])
        wb.save(product_filepath)
    
    # Materials data
    materials_filepath = 'data/materials.json'
    if not os.path.exists(materials_filepath):
        with open(materials_filepath, 'w') as f:
            json.dump([], f)
    
    # Study resources
    study_filepath = 'data/study.json'
    if not os.path.exists(study_filepath):
        with open(study_filepath, 'w') as f:
            json.dump([], f)
    
    # Settings
    settings_filepath = 'data/settings.json'
    if not os.path.exists(settings_filepath):
        with open(settings_filepath, 'w') as f:
            json.dump({'language': 'EN', 'theme': 'light'}, f)
    
    return {
        'products': product_filepath,
        'materials': materials_filepath,
        'study': study_filepath,
        'settings': settings_filepath
    }

# AI Functions with enhanced Tamil support
def generate_description(product_name, lang):
    if lang == 'TA':
        return f"உயர்தர {product_name}, பாரம்பரிய கைவினைத் திறனால் உருவாக்கப்பட்டது"
    return f"Handcrafted {product_name} made with traditional techniques"

def predict_category(product_name):
    # Tamil product recognition
    tamil_keywords = {
        'புடவை': 'Textiles',
        'சேலை': 'Textiles',
        'மட்பாண்டம்': 'Pottery',
        'விளக்கு': 'Pottery',
        'செம்பு': 'Metalwork',
        'வெண்கலம்': 'Metalwork',
        'நகை': 'Jewelry',
        'சிற்பம்': 'Sculpture',
        'மரம்': 'Woodwork',
        'பட்டு': 'Silk'
    }
    
    # Check for Tamil keywords
    for keyword, category in tamil_keywords.items():
        if keyword in product_name:
            return category
    
    # English fallback
    if 'silk' in product_name.lower() or 'saree' in product_name.lower():
        return 'Textiles'
    elif 'pot' in product_name.lower() or 'vase' in product_name.lower() or 'lamp' in product_name.lower():
        return 'Pottery'
    elif 'metal' in product_name.lower() or 'brass' in product_name.lower() or 'bronze' in product_name.lower():
        return 'Metalwork'
    elif 'jewelry' in product_name.lower() or 'necklace' in product_name.lower() or 'ring' in product_name.lower():
        return 'Jewelry'
    elif 'wood' in product_name.lower() or 'carving' in product_name.lower():
        return 'Woodwork'
    elif 'painting' in product_name.lower() or 'art' in product_name.lower():
        return 'Art'
    return 'Handicrafts'

# Tamil number words to digits mapping
tamil_numbers_dict = {
    'ஒன்று': 1, 'இரண்டு': 2, 'மூன்று': 3, 'நான்கு': 4, 'ஐந்து': 5,
    'ஆறு': 6, 'ஏழு': 7, 'எட்டு': 8, 'ஒன்பது': 9, 'பத்து': 10,
    'பதினொன்று': 11, 'பன்னிரண்டு': 12, 'பதிமூன்று': 13, 'பதினான்கு': 14, 'பதினைந்து': 15,
    'பதினாறு': 16, 'பதினேழு': 17, 'பதினெட்டு': 18, 'பத்தொன்பது': 19, 'இருபது': 20,
    'இருபத்தி ஒன்று': 21, 'இருபத்தி இரண்டு': 22, 'முப்பது': 30, 'நாற்பது': 40, 'ஐம்பது': 50,
    'அறுபது': 60, 'எழுபது': 70, 'எண்பது': 80, 'தொண்ணூறு': 90, 'நூறு': 100,
    'இருநூறு': 200, 'முன்னூறு': 300, 'நாநூறு': 400, 'ஐநூறு': 500, 'ஆறுநூறு': 600,
    'எழுநூறு': 700, 'எண்ணூறு': 800, 'தொள்ளாயிரம்': 900, 'ஆயிரம்': 1000
}

def extract_quantity(input_text):
    # First try to find Tamil number words
    for word, number in tamil_numbers_dict.items():
        if word in input_text:
            return number
    
    # Then look for digits
    numbers = re.findall(r'\d+', input_text)
    return int(numbers[0]) if numbers else 1

def extract_price(input_text):
    # Try to find Tamil number words with currency context
    for word, number in tamil_numbers_dict.items():
        if word in input_text and ('ரூபாய்' in input_text or 'ரூ' in input_text):
            return number
    
    # Look for currency patterns
    matches = re.findall(r'(\d+)\s*(ரூபாய்|ரூ|rupees?|rs|₹)', input_text, re.IGNORECASE)
    if matches:
        return int(matches[-1][0])
    
    # If no currency pattern, look for any numbers
    numbers = re.findall(r'\d+', input_text)
    return int(numbers[-1]) if numbers else 1000

# Existing product routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/add_product', methods=['POST'])
def add_product():
    try:
        files = init_data_files()
        wb = openpyxl.load_workbook(files['products'])
        ws = wb.active
        
        data = request.json
        product_name = data['name'].strip()
        quantity = int(data['quantity'])
        price = int(data['price'])
        
        # Validate inputs
        if not product_name or quantity <= 0 or price <= 0:
            return jsonify(success=False, error="Invalid input values"), 400
        
        category = predict_category(product_name)
        desc_en = generate_description(product_name, 'EN')
        desc_ta = generate_description(product_name, 'TA')
        
        # Generate unique ID
        new_id = str(uuid4())
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        ws.append([
            new_id, 
            product_name, 
            category, 
            desc_en, 
            desc_ta, 
            quantity, 
            price, 
            timestamp
        ])
        
        wb.save(files['products'])
        return jsonify(success=True, id=new_id)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route('/get_products')
def get_products():
    try:
        files = init_data_files()
        wb = openpyxl.load_workbook(files['products'])
        ws = wb.active
        
        products = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Skip empty rows
                products.append({
                    'id': row[0],
                    'name': row[1],
                    'category': row[2],
                    'desc_en': row[3],
                    'desc_ta': row[4],
                    'quantity': row[5],
                    'price': row[6],
                    'timestamp': row[7]
                })
        
        return jsonify(products)
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route('/delete_product/<string:product_id>', methods=['DELETE'])
def delete_product(product_id):
    try:
        files = init_data_files()
        wb = openpyxl.load_workbook(files['products'])
        ws = wb.active
        
        row_index = None
        for idx, row in enumerate(ws.iter_rows(min_row=2, min_col=1, max_col=1), 2):
            if row[0].value == product_id:
                row_index = idx
                break
        
        if row_index:
            ws.delete_rows(row_index)
            wb.save(files['products'])
            return jsonify(success=True)
        return jsonify(success=False, error="Product not found"), 404
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route('/download_excel')
def download_excel():
    files = init_data_files()
    return send_file(files['products'], as_attachment=True)

@app.route('/voice_command', methods=['POST'])
def handle_voice_command():
    try:
        data = request.json
        transcript = data['transcript']
        lang = data.get('lang', 'EN')
        
        # Process the voice command
        command = transcript.lower()
        
        # Simple command matching
        if 'add new product' in command or 'புதிய பொருள் சேர்' in command:
            return jsonify(action='open_add_product')
        elif 'show low stock' in command or 'குறைந்த சரக்கு காட்டு' in command:
            return jsonify(action='filter_low_stock')
        elif 'open sales report' in command or 'விற்பனை அறிக்கை திற' in command:
            return jsonify(action='open_sales_report')
        elif 'go to materials' in command or 'பொருட்கள் பிரிவுக்கு செல்' in command:
            return jsonify(action='navigate_materials')
        elif 'show analysis report' in command or 'ஆய்வு அறிக்கை காட்டு' in command:
            return jsonify(action='open_analysis')
        elif 'show dashboard' in command or 'டாஷ்போர்டு காட்டு' in command:
            return jsonify(action='open_dashboard')
        elif 'show ai insights' in command or 'ai பரிந்துரைகள் காட்டு' in command:
            return jsonify(action='open_insights')
        elif 'open settings' in command or 'அமைப்புகள் திற' in command:
            return jsonify(action='open_settings')
        else:
            # Try to extract product details
            quantity = extract_quantity(transcript)
            price = extract_price(transcript)
            
            # Extract product name by removing numbers and currency words
            product_name = re.sub(r'\d+', '', transcript)
            for word in ['ரூபாய்', 'rupees', 'ரூ', 'rs', '₹', 'price', 'விலை', 'quantity', 'எண்ணிக்கை']:
                product_name = product_name.replace(word, '')
            
            return jsonify(
                action='add_product',
                product_name=product_name.strip(),
                quantity=quantity,
                price=price
            )
            
    except Exception as e:
        return jsonify(error=str(e)), 500

if __name__ == '__main__':
    app.run(debug=True)